import openpyxl as xl
from openpyxl.chart import Reference, BarChart


def workbook(workbook_name):
    if '.xlsx' in workbook_name:
        pass
    else:
        workbook_name += '.xlsx'

    wb = xl.load_workbook(workbook_name)
    sheet1 = wb['Sheet1']

    actual_rows = sum(1 for row in sheet1.iter_rows(values_only=True) if any(row))
    actual_cols = sum(1 for col in sheet1.iter_cols(values_only=True) if any(col))

    sheet1.cell(1, actual_cols + 1).value = "new Price"

    for row in range(2, actual_rows + 1):
        current_sales_price = sheet1.cell(row, actual_cols).value
        new_sales_price = current_sales_price * 0.9
        new_sales_cell = sheet1.cell(row, actual_cols + 1)
        new_sales_cell.value = new_sales_price

    # values need to draw graph
    value_refs = Reference(sheet1,
                           min_row=2,
                           max_row=actual_rows,
                           min_col=actual_cols+1,
                           max_col=actual_cols+1)
    # draw bar chat
    bar_chat01 = BarChart()
    bar_chat01.add_data(value_refs)
    sheet1.add_chart(bar_chat01, 'f5')

    wb.save('new_workbook.xlsx')


workbook('transaction')
